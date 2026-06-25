import os
import json
import uuid
import re
import csv
import pdfplumber
import openpyxl
from io import BytesIO, StringIO
from typing import List
from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from openai import OpenAI
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.config import settings
from app.database import get_db
from app.models.document import Document
from app.models.employee import Employee
from app.routers.auth import get_current_employee
from app.services.r2_service import download_file_from_r2, list_r2_objects

def is_structured_schedule_csv(file_text: str) -> bool:
    if not file_text:
        return False
    lines = [line.strip() for line in file_text.split('\n') if line.strip()]
    if not lines:
        return False
    first_line = lines[0]
    delimiters = [',', '\t', ';']
    for delim in delimiters:
        parts = [p.strip().lower() for p in first_line.split(delim)]
        has_order = any('주문번호' in p or 'order_num' in p or 'order_id' in p for p in parts)
        has_task = any('작업id' in p or 'task_id' in p or '작업 id' in p or '작업코드' in p for p in parts)
        has_start = any('시작일' in p or 'start_date' in p or '시작 일' in p or '시작일시' in p for p in parts)
        has_end = any('종료일' in p or 'end_date' in p or '종료 일' in p or '종료일시' in p for p in parts)
        if has_order and has_task and has_start and has_end:
            return True
    return False

def parse_schedule_csv_directly(
    file_text: str,
    db: Session,
    employee_map: dict,
    equipment_map: dict,
    order_map: dict,
    valid_tasks: set
) -> int:
    lines = [line.strip() for line in file_text.split('\n') if line.strip()]
    if not lines:
        return 0
    first_line = lines[0]
    delimiter = ','
    for d in ['\t', ';']:
        if d in first_line:
            delimiter = d
            break
            
    f = StringIO(file_text)
    reader = csv.DictReader(f, delimiter=delimiter)
    
    def find_column(headers, candidates):
        headers_norm = [h.strip().lower() for h in headers]
        for cand in candidates:
            cand_norm = cand.lower().strip()
            if cand_norm in headers_norm:
                idx = headers_norm.index(cand_norm)
                return headers[idx]
        for cand in candidates:
            cand_norm = cand.lower().strip()
            for i, h in enumerate(headers_norm):
                if cand_norm in h or h in cand_norm:
                    return headers[i]
        return None

    headers = reader.fieldnames
    if not headers:
        return 0
        
    col_order_num = find_column(headers, ["주문번호", "order_num", "order_id"])
    col_task_id = find_column(headers, ["작업ID", "작업id", "task_id", "작업 id"])
    col_start_date = find_column(headers, ["시작일", "start_date", "시작 일", "시작일시"])
    col_end_date = find_column(headers, ["종료일", "end_date", "종료 일", "종료일시"])
    col_factory = find_column(headers, ["공장동", "factory", "공장"])
    col_workers = find_column(headers, ["배정직원", "workers", "employees", "작업자"])
    col_product = find_column(headers, ["제품명", "product_name", "제품"])
    col_count = find_column(headers, ["수량", "order_count", "주문수량"])
    col_due_date = find_column(headers, ["납기일", "due_date", "납기"])
    col_status = find_column(headers, ["납기상태", "order_status", "상태"])
    col_equip = find_column(headers, ["필요장비", "equipments", "equipment", "장비"])
    
    saved_count = 0
    
    # Refresh order_map from DB
    orders_db = db.execute(text("SELECT order_id, order_num FROM orders")).mappings().all()
    for o in orders_db:
        ord_id = o['order_id']
        ord_num = o['order_num']
        order_map[ord_id.lower().strip()] = ord_id
        order_map[ord_num.lower().strip()] = ord_id
        
    for row in reader:
        order_num = row.get(col_order_num, "").strip() if col_order_num else ""
        task_id = row.get(col_task_id, "").strip() if col_task_id else ""
        start_date_str = row.get(col_start_date, "").strip() if col_start_date else ""
        end_date_str = row.get(col_end_date, "").strip() if col_end_date else ""
        
        if not order_num or not task_id or not start_date_str or not end_date_str:
            continue
            
        cleaned_task = task_id.upper().strip()
        if cleaned_task in valid_tasks:
            task_id = cleaned_task
        elif task_id not in valid_tasks:
            print(f"⚠️ Direct Parser: Invalid task_id ({task_id}), skipping row.")
            continue
            
        # Resolve order_id
        num_key = order_num.lower().strip()
        if num_key in order_map:
            order_id = order_map[num_key]
        else:
            order_id = f"ord_new_{uuid.uuid4().hex[:8]}"
            order_map[num_key] = order_id
            
        product_name = row.get(col_product, "UNKNOWN").strip() if col_product else "UNKNOWN"
        order_count_val = row.get(col_count, "0").strip() if col_count else "0"
        m_cnt = re.search(r'\d+', order_count_val)
        order_count = int(m_cnt.group(0)) if m_cnt else 0
        due_date = row.get(col_due_date, "2026-07-31").strip() if col_due_date else "2026-07-31"
        order_status = row.get(col_status, "COMPLETED").strip() if col_status else "COMPLETED"
        
        if "완료" in order_status or "COMPLETED" in order_status.upper():
            order_status = "COMPLETED"
        elif "진행" in order_status or "IN_PROGRESS" in order_status.upper():
            order_status = "IN_PROGRESS"
        else:
            order_status = "PENDING"
            
        try:
            with db.begin_nested():
                db.execute(
                    text("""
                        INSERT INTO orders (order_id, order_num, product_name, order_count, due_date, order_status)
                        VALUES (:order_id, :order_num, :product_name, :order_count, :due_date, :order_status)
                        ON CONFLICT (order_id) DO UPDATE SET
                            order_num = EXCLUDED.order_num,
                            product_name = EXCLUDED.product_name,
                            order_count = EXCLUDED.order_count,
                            due_date = EXCLUDED.due_date,
                            order_status = EXCLUDED.order_status
                    """),
                    {
                        "order_id": order_id,
                        "order_num": order_num,
                        "product_name": product_name,
                        "order_count": order_count,
                        "due_date": due_date,
                        "order_status": order_status
                    }
                )
        except Exception as e:
            print(f"❌ Direct Parser: Order insert error: {str(e)}")
            continue
            
        if len(start_date_str) == 10:
            start_date_str += " 09:00:00"
        if len(end_date_str) == 10:
            end_date_str += " 18:00:00"
            
        factory = row.get(col_factory, "A동").strip() if col_factory else "A동"
        if not factory:
            factory = "A동"
            
        # Parse workers
        workers_str = row.get(col_workers, "") if col_workers else ""
        workers = []
        if workers_str:
            emp_finds = re.findall(r'([eE][mM][pP]\d+)', workers_str)
            for emp_id in emp_finds:
                key = emp_id.lower().strip()
                if key in employee_map:
                    workers.append(employee_map[key])
            tokens = re.split(r'[^a-zA-Z0-9가-힣]+', workers_str)
            for token in tokens:
                key = token.lower().strip()
                if key in employee_map:
                    workers.append(employee_map[key])
            workers = list(set(workers))
            
        # Parse equipment
        equip_str = row.get(col_equip, "") if col_equip else ""
        equip_ids = []
        if equip_str:
            tokens = re.split(r'[;,/]+', equip_str)
            for token in tokens:
                token = token.strip()
                if not token:
                    continue
                eq_id = None
                key = token.lower()
                if key in equipment_map:
                    eq_id = equipment_map[key]
                else:
                    m = re.match(r'^장비\s*(\d+)$', token)
                    if m:
                        eq_num = int(m.group(1))
                        candidate_id = f"EQ{eq_num:03d}"
                        if candidate_id.lower() in equipment_map:
                            eq_id = equipment_map[candidate_id.lower()]
                    else:
                        m2 = re.match(r'^(?:[eE][qQ])?\s*(\d+)$', token)
                        if m2:
                            eq_num = int(m2.group(1))
                            candidate_id = f"EQ{eq_num:03d}"
                            if candidate_id.lower() in equipment_map:
                                eq_id = equipment_map[candidate_id.lower()]
                if eq_id:
                    equip_ids.append(eq_id)
            equip_ids = list(set(equip_ids))
            
        # Insert required equipments
        for eq_id in equip_ids:
            try:
                with db.begin_nested():
                    db.execute(
                        text("""
                            INSERT INTO required_equipments (task_id, eq_id)
                            VALUES (:task_id, :eq_id)
                            ON CONFLICT (task_id, eq_id) DO NOTHING
                        """),
                        {"task_id": task_id, "eq_id": eq_id}
                    )
            except Exception as e:
                print(f"❌ Direct Parser: Required equipment insert error: {str(e)}")
                
        # Insert/Update schedules
        existing = db.execute(
            text("SELECT id FROM schedules WHERE task_id = :task_id AND order_id = :order_id"),
            {"task_id": task_id, "order_id": order_id}
        ).mappings().first()
        
        if existing:
            schedule_id = existing["id"]
            try:
                with db.begin_nested():
                    db.execute(
                        text("""
                            UPDATE schedules 
                            SET start_date = :start_date, end_date = :end_date, factory = :factory
                            WHERE id = :id AND task_id = :task_id AND order_id = :order_id
                        """),
                        {
                            "id": schedule_id,
                            "task_id": task_id,
                            "order_id": order_id,
                            "start_date": start_date_str,
                            "end_date": end_date_str,
                            "factory": factory,
                        }
                    )
                    db.execute(
                        text("""
                            DELETE FROM schedule_assignments 
                            WHERE id = :id AND task_id = :task_id AND order_id = :order_id
                        """),
                        {
                            "id": schedule_id,
                            "task_id": task_id,
                            "order_id": order_id
                        }
                    )
            except Exception as e:
                print(f"❌ Direct Parser: Schedule update error: {str(e)}")
                continue
        else:
            schedule_id = f"sch_{uuid.uuid4().hex[:8]}"
            try:
                with db.begin_nested():
                    db.execute(
                        text("""
                            INSERT INTO schedules (id, task_id, order_id, start_date, end_date, factory)
                            VALUES (:id, :task_id, :order_id, :start_date, :end_date, :factory)
                        """),
                        {
                            "id": schedule_id,
                            "task_id": task_id,
                            "order_id": order_id,
                            "start_date": start_date_str,
                            "end_date": end_date_str,
                            "factory": factory,
                        }
                    )
            except Exception as e:
                print(f"❌ Direct Parser: Schedule insert error: {str(e)}")
                continue
                
        has_at_least_one_worker = False
        for worker_id in workers:
            try:
                with db.begin_nested():
                    db.execute(
                        text("""
                            INSERT INTO schedule_assignments (id, user_id, task_id, order_id)
                            VALUES (:id, :user_id, :task_id, :order_id)
                            ON CONFLICT (id, user_id, task_id, order_id) DO NOTHING
                        """),
                        {
                            "id": schedule_id,
                            "user_id": worker_id,
                            "task_id": task_id,
                            "order_id": order_id
                        }
                    )
                has_at_least_one_worker = True
            except Exception as e:
                print(f"❌ Direct Parser: Assignment insert error: {str(e)}")
                
        if has_at_least_one_worker:
            saved_count += 1
            
    db.commit()
    return saved_count

router = APIRouter(prefix="/api/schedule", tags=["Schedule Generator"])
client = OpenAI(api_key=settings.OPENAI_API_KEY)

class GenerateRequest(BaseModel):
    file_ids: List[str]

def extract_text_from_pdf(file_path: str) -> str:
    text_content = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text_content += page.extract_text() or ""
    return text_content

def extract_text_from_excel(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    text_content = ""
    for sheet in wb.worksheets:
        text_content += f"\n[시트: {sheet.title}]\n"
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
            if row_text.strip():
                text_content += row_text + "\n"
    return text_content

def extract_text_from_bytes(file_bytes: bytes, filename: str) -> str:
    """바이트 데이터에서 텍스트 추출"""
    ext = os.path.splitext(filename)[1].lower()
    
    if ext == ".pdf":
        text_content = ""
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                text_content += page.extract_text() or ""
        return text_content
    
    elif ext in [".xlsx", ".xls"]:
        text_content = ""
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        for sheet in wb.worksheets:
            text_content += f"\n[시트: {sheet.title}]\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text_content += row_text + "\n"
        return text_content
    
    elif ext == ".csv":
        return file_bytes.decode('utf-8', errors='ignore')
    
    elif ext in [".txt"]:
        return file_bytes.decode('utf-8', errors='ignore')
    
    else:
        raise ValueError(f"지원하지 않는 형식: {filename}")

def chunk_document_text(all_text: str, chunk_size: int = 100) -> list[str]:
    """텍스트를 라인 단위로 chunk_size만큼 나눔. 파일 및 시트 헤더는 각 chunk에 유지함."""
    lines = [line for line in all_text.split("\n") if line.strip()]
    chunks = []
    
    # We will keep track of the current file/sheet headers
    current_file_header = ""
    current_sheet_header = ""
    current_column_header = ""
    
    current_chunk = []
    
    for line in lines:
        stripped = line.strip()
        
        # Check if this line is a file or sheet header
        if stripped.startswith("[파일:"):
            # Save the current chunk if it exists before switching files
            if current_chunk:
                header_text = "\n".join(filter(None, [current_file_header, current_sheet_header, current_column_header]))
                chunks.append(header_text + "\n" + "\n".join(current_chunk))
                current_chunk = []
            current_file_header = line
            current_sheet_header = ""
            current_column_header = ""
            continue
            
        elif stripped.startswith("[시트:"):
            # Save the current chunk if it exists before switching sheets
            if current_chunk:
                header_text = "\n".join(filter(None, [current_file_header, current_sheet_header, current_column_header]))
                chunks.append(header_text + "\n" + "\n".join(current_chunk))
                current_chunk = []
            current_sheet_header = line
            current_column_header = ""
            continue
            
        # If we don't have a column header yet, the first non-header line must be the column header!
        elif not current_column_header:
            current_column_header = line
            continue
            
        # Otherwise, it's a regular data row
        current_chunk.append(line)
        
        if len(current_chunk) >= chunk_size:
            header_text = "\n".join(filter(None, [current_file_header, current_sheet_header, current_column_header]))
            chunks.append(header_text + "\n" + "\n".join(current_chunk))
            current_chunk = []
            
    if current_chunk:
        header_text = "\n".join(filter(None, [current_file_header, current_sheet_header, current_column_header]))
        chunks.append(header_text + "\n" + "\n".join(current_chunk))
        
    return chunks


@router.post("/generate", summary="업로드된 파일 기반 일정 생성 및 저장")
async def generate_schedule(
    body: GenerateRequest,
    db: Session = Depends(get_db),
    current_emp: Employee = Depends(get_current_employee)
):
    tasks = db.execute(text("SELECT task_id, task_name FROM task")).mappings().all()
    task_list = "\n".join([f"- {t['task_id']}: {t['task_name']}" for t in tasks])

    orders = db.execute(text("SELECT order_id, order_num, product_name FROM orders")).mappings().all()
    order_list = "\n".join([f"- {o['order_id']}: {o['order_num']} ({o['product_name']})" for o in orders])

    employees = db.execute(text("SELECT emp_id, login_id, emp_name, emp_role FROM employees")).mappings().all()
    employee_list = "\n".join([f"- {e['emp_id']}: {e['emp_name']} ({e['emp_role']})" for e in employees])

    equipments = db.execute(text("SELECT eq_id, eq_name FROM equipments")).mappings().all()
    equipment_list = "\n".join([f"- {eq['eq_id']}: {eq['eq_name']}" for eq in equipments])

    valid_tasks = {t['task_id'] for t in tasks}

    # Map name/login_id/emp_id to emp_id
    employee_map = {}
    for e in employees:
        emp_id = e['emp_id']
        emp_name = e['emp_name']
        login_id = e['login_id']
        employee_map[emp_id.lower().strip()] = emp_id
        employee_map[emp_name.lower().strip()] = emp_id
        if login_id:
            employee_map[login_id.lower().strip()] = emp_id

    # Map name/eq_id to eq_id
    equipment_map = {}
    for eq in equipments:
        eq_id = eq['eq_id']
        eq_name = eq['eq_name']
        equipment_map[eq_id.lower().strip()] = eq_id
        equipment_map[eq_name.lower().strip()] = eq_id

    # Map order_id/order_num to order_id
    order_map = {}
    for o in orders:
        ord_id = o['order_id']
        ord_num = o['order_num']
        order_map[ord_id.lower().strip()] = ord_id
        order_map[ord_num.lower().strip()] = ord_id

    all_text = ""
    direct_parsed_schedules = 0
    direct_parsed_any = False

    for file_id in body.file_ids:
        doc = db.query(Document).filter(
            Document.file_id == file_id,
            Document.uploader == current_emp.emp_id
        ).first()

        if not doc:
            raise HTTPException(status_code=404, detail=f"파일을 찾을 수 없습니다: {file_id}")

        if not os.path.exists(doc.file_path):
            raise HTTPException(status_code=404, detail=f"파일이 서버에 없습니다: {doc.file_name}")

        ext = doc.file_extension.lower()
        if ext == ".pdf":
            file_text = extract_text_from_pdf(doc.file_path)
        elif ext in [".xlsx", ".xls"]:
            file_text = extract_text_from_excel(doc.file_path)
        elif ext == ".csv":
            try:
                with open(doc.file_path, "r", encoding="utf-8") as f:
                    file_text = f.read()
            except UnicodeDecodeError:
                with open(doc.file_path, "r", encoding="cp949", errors="ignore") as f:
                    file_text = f.read()
        elif ext == ".txt":
            try:
                with open(doc.file_path, "r", encoding="utf-8") as f:
                    file_text = f.read()
            except UnicodeDecodeError:
                with open(doc.file_path, "r", encoding="cp949", errors="ignore") as f:
                    file_text = f.read()
        else:
            raise HTTPException(status_code=400, detail=f"{doc.file_name}: PDF, Excel, CSV 또는 TXT만 지원합니다.")

        # Check if it is structured schedule CSV
        if is_structured_schedule_csv(file_text):
            print(f"ℹ️ Detected structured CSV '{doc.file_name}'. Parsing directly.")
            direct_count = parse_schedule_csv_directly(
                file_text,
                db=db,
                employee_map=employee_map,
                equipment_map=equipment_map,
                order_map=order_map,
                valid_tasks=valid_tasks
            )
            direct_parsed_schedules += direct_count
            direct_parsed_any = True
        else:
            print(f"ℹ️ Skipping non-schedule file: {doc.file_name}")
            continue

    llm_parsed_schedules = 0
    combined_result = {
        "orders": [],
        "schedules": [],
        "required_equipments": [],
        "summaries": []
    }

    if all_text.strip():
        # Chunking
        all_chunks = chunk_document_text(all_text, chunk_size=100)
        print(f"ℹ️ Split document into {len(all_chunks)} chunks for LLM processing.")

        for idx, chunk in enumerate(all_chunks):
            prompt = f"""다음은 제조 공장의 생산 관련 문서들입니다. 이 문서들을 분석하여 생산 관련 데이터(주문, 일정, 작업자 배정, 필요 장비)를 JSON 형식으로 생성해주세요.

문서 내용:
{chunk}

사용 가능한 task 목록 (반드시 아래 task_id 중에서 선택):
{task_list}

사용 가능한 기존 order 목록 (문서에 있는 주문이 기존 목록에 있으면 해당 order_id를 사용하고, 없으면 새로운 order_id를 생성하여 orders 목록에 추가하세요):
{order_list}

사용 가능한 직원 목록 (반드시 아래 emp_id 또는 직원 이름 중에서 선택하여 일정에 배정):
{employee_list}

사용 가능한 장비 목록 (반드시 아래 eq_id 또는 장비 이름 중에서 선택하여 작업에 매핑):
{equipment_list}

다음 형식으로 응답해주세요:
{{
  "orders": [
    {{
      "order_id": "ord_001", // 기존 order_id 또는 신규 생성된 ID (예: ord_new_001)
      "order_num": "PO001",
      "product_name": "DRAM-8G",
      "order_count": 2500,
      "due_date": "2026-07-31",
      "order_status": "COMPLETED" // COMPLETED, IN_PROGRESS, PENDING 등
    }}
  ],
  "schedules": [
    {{
      "task_id": "tsk_001",
      "order_id": "ord_001",
      "start_date": "2026-07-01 09:00:00",
      "end_date": "2026-07-02 18:00:00",
      "factory": "A동", // 공장동 (예: A동, B동 등)
      "workers": ["emp001", "emp002"] // 해당 일정에 배정될 직원 ID 또는 직원 이름 목록 (최소 1명 이상 배정)
    }}
  ],
  "required_equipments": [
    {{
      "task_id": "tsk_001",
      "eq_id": "eq_001" // 해당 task에 필요한 장비 ID 또는 장비 이름
    }}
  ],
  "summary": "일정 요약 설명"
}}

주의사항:
1. schedules의 order_id는 반드시 orders 목록에 정의되어 있거나 기존 order 목록에 있는 ID/주문번호여야 합니다.
2. schedules의 task_id는 반드시 사용 가능한 task 목록에 있는 ID여야 합니다.
3. schedules의 workers는 반드시 사용 가능한 직원 목록의 emp_id 또는 직원 이름에서 선택해야 합니다. (최소 1명 이상 배정해야 일정 조회 쿼리에서 누락되지 않습니다)
4. required_equipments의 task_id 및 eq_id는 각각 제공된 task 목록과 장비 목록에 있는 ID/이름이어야 합니다.
5. JSON만 응답하고 다른 텍스트는 포함하지 마세요."""

            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3,
            )

            result_text = response.choices[0].message.content.strip()
            result_text = result_text.replace("```json", "").replace("```", "").strip()

            try:
                chunk_result = json.loads(result_text)
                combined_result["orders"].extend(chunk_result.get("orders", []))
                combined_result["schedules"].extend(chunk_result.get("schedules", []))
                combined_result["required_equipments"].extend(chunk_result.get("required_equipments", []))
                if chunk_result.get("summary"):
                    combined_result["summaries"].append(chunk_result["summary"])
            except Exception as e:
                print(f"❌ Error processing chunk {idx}: {str(e)}")
                continue

        # Deduplicate orders and required equipments in Python
        unique_orders = []
        seen_order_ids = set()
        for o in combined_result["orders"]:
            oid = o.get("order_id")
            if oid and oid not in seen_order_ids:
                seen_order_ids.add(oid)
                unique_orders.append(o)

        unique_required = []
        seen_req = set()
        for r in combined_result["required_equipments"]:
            t_id = r.get("task_id")
            eq_val = r.get("eq_id")
            if t_id and eq_val:
                key = (t_id, str(eq_val).lower().strip())
                if key not in seen_req:
                    seen_req.add(key)
                    unique_required.append(r)

        # 1. Insert/Update Orders
        resolved_orders = {}
        for o in unique_orders:
            llm_order_id = o.get("order_id")
            order_num = o.get("order_num", "").strip()
            if not llm_order_id:
                continue
                
            num_key = order_num.lower().strip()
            id_key = llm_order_id.lower().strip()
            
            # Resolve to existing order_id if order_num or order_id matches
            if num_key in order_map:
                correct_id = order_map[num_key]
            elif id_key in order_map:
                correct_id = order_map[id_key]
            else:
                correct_id = llm_order_id
                order_map[id_key] = correct_id
                if order_num:
                    order_map[num_key] = correct_id
                    
            resolved_orders[llm_order_id] = correct_id
            
            try:
                with db.begin_nested():
                    db.execute(
                        text("""
                            INSERT INTO orders (order_id, order_num, product_name, order_count, due_date, order_status)
                            VALUES (:order_id, :order_num, :product_name, :order_count, :due_date, :order_status)
                            ON CONFLICT (order_id) DO UPDATE SET
                                order_num = EXCLUDED.order_num,
                                product_name = EXCLUDED.product_name,
                                order_count = EXCLUDED.order_count,
                                due_date = EXCLUDED.due_date,
                                order_status = EXCLUDED.order_status
                        """),
                        {
                            "order_id": correct_id,
                            "order_num": order_num or "UNKNOWN",
                            "product_name": o.get("product_name", "UNKNOWN"),
                            "order_count": int(o.get("order_count", 0)),
                            "due_date": o.get("due_date", "2026-07-01"),
                            "order_status": o.get("order_status", "PENDING"),
                        }
                    )
            except Exception as e:
                print(f"❌ Order insert/update error: {str(e)}")
                continue

        # 2. Insert Required Equipments (with mathematical symbol resolution)
        for req in unique_required:
            task_id = req.get("task_id")
            eq_val = req.get("eq_id")
            if not task_id or not eq_val:
                continue
            if task_id not in valid_tasks:
                print(f"⚠️ Invalid task_id ({task_id}) for required_equipments, skipping.")
                continue
            eq_key = str(eq_val).lower().strip()
            eq_id = None
            if eq_key in equipment_map:
                eq_id = equipment_map[eq_key]
            else:
                m = re.match(r'^장비\s*(\d+)$', str(eq_val).strip())
                if m:
                    candidate = f"EQ{int(m.group(1)):03d}"
                    if candidate.lower() in equipment_map:
                        eq_id = equipment_map[candidate.lower()]
                else:
                    m2 = re.match(r'^(?:[eE][qQ])?\s*(\d+)$', str(eq_val).strip())
                    if m2:
                        candidate = f"EQ{int(m2.group(1)):03d}"
                        if candidate.lower() in equipment_map:
                            eq_id = equipment_map[candidate.lower()]
            if not eq_id:
                print(f"⚠️ Equipment '{eq_val}' not resolved, skipping required_equipment mapping.")
                continue
            try:
                with db.begin_nested():
                    db.execute(
                        text("""
                            INSERT INTO required_equipments (task_id, eq_id)
                            VALUES (:task_id, :eq_id)
                            ON CONFLICT (task_id, eq_id) DO NOTHING
                        """),
                        {
                            "task_id": task_id,
                            "eq_id": eq_id
                        }
                    )
            except Exception as e:
                print(f"❌ Required equipment insert error: {str(e)}")
                continue

        # 3. Insert/Update Schedules
        saved = []
        for s in combined_result["schedules"]:
            task_id = s.get("task_id")
            order_val = s.get("order_id")
            if not task_id or not order_val:
                continue
            if task_id not in valid_tasks:
                print(f"⚠️ Invalid task_id ({task_id}) for schedule, skipping.")
                continue

            # Resolve order_id
            if order_val in resolved_orders:
                order_id = resolved_orders[order_val]
            else:
                order_key = str(order_val).lower().strip()
                if order_key in order_map:
                    order_id = order_map[order_key]
                else:
                    print(f"⚠️ Order_id ({order_val}) not found in orders database or generated list, skipping schedule.")
                    continue

            # Check if schedule already exists for this task and order to prevent duplicates
            existing = db.execute(
                text("SELECT id FROM schedules WHERE task_id = :task_id AND order_id = :order_id"),
                {"task_id": task_id, "order_id": order_id}
            ).mappings().first()

            if existing:
                schedule_id = existing["id"]
                try:
                    with db.begin_nested():
                        db.execute(
                            text("""
                                UPDATE schedules 
                                SET start_date = :start_date, end_date = :end_date, factory = :factory
                                WHERE id = :id AND task_id = :task_id AND order_id = :order_id
                            """),
                            {
                                "id": schedule_id,
                                "task_id": task_id,
                                "order_id": order_id,
                                "start_date": s["start_date"],
                                "end_date": s["end_date"],
                                "factory": s.get("factory", "A동"),
                            }
                        )
                        # Clean up old assignments for this schedule to prevent duplicates
                        db.execute(
                            text("""
                                DELETE FROM schedule_assignments 
                                WHERE id = :id AND task_id = :task_id AND order_id = :order_id
                            """),
                            {
                                "id": schedule_id,
                                "task_id": task_id,
                                "order_id": order_id
                            }
                        )
                except Exception as e:
                    print(f"❌ Schedule update error: {str(e)}")
                    continue
            else:
                schedule_id = f"sch_{uuid.uuid4().hex[:8]}"
                try:
                    with db.begin_nested():
                        db.execute(
                            text("""
                                INSERT INTO schedules (id, task_id, order_id, start_date, end_date, factory)
                                VALUES (:id, :task_id, :order_id, :start_date, :end_date, :factory)
                            """),
                            {
                                "id": schedule_id,
                                "task_id": task_id,
                                "order_id": order_id,
                                "start_date": s["start_date"],
                                "end_date": s["end_date"],
                                "factory": s.get("factory", "A동"),
                            }
                        )
                except Exception as e:
                    print(f"❌ Schedule insert error: {str(e)}")
                    continue

            # Map workers to valid employee IDs
            workers = []
            for w in s.get("workers", []):
                w_key = str(w).lower().strip()
                if w_key in employee_map:
                    workers.append(employee_map[w_key])
                else:
                    print(f"⚠️ Worker value '{w}' not resolved to an employee.")

            # De-duplicate workers in Python
            workers = list(set(workers))
            saved.append((schedule_id, task_id, order_id, workers))

        # 4. Insert Schedule Assignments
        for schedule_id, task_id, order_id, workers in saved:
            has_at_least_one_worker = False
            for worker_id in workers:
                try:
                    with db.begin_nested():
                        db.execute(
                            text("""
                                INSERT INTO schedule_assignments (id, user_id, task_id, order_id)
                                VALUES (:id, :user_id, :task_id, :order_id)
                                ON CONFLICT (id, user_id, task_id, order_id) DO NOTHING
                            """),
                            {
                                "id": schedule_id,
                                "user_id": worker_id,
                                "task_id": task_id,
                                "order_id": order_id
                            }
                        )
                    has_at_least_one_worker = True
                except Exception as e:
                    print(f"❌ Assignment insert error: {str(e)}")
                    continue
            if has_at_least_one_worker:
                llm_parsed_schedules += 1

        db.commit()

    total_saved = direct_parsed_schedules + llm_parsed_schedules
    summary_messages = []
    if direct_parsed_any:
        summary_messages.append(f"구조화된 CSV 파일에서 {direct_parsed_schedules}개 일정을 직접 파싱하여 저장했습니다.")
    if all_text.strip():
        summary_messages.append("\n".join(combined_result["summaries"]))

    return {
        "message": f"일정 {total_saved}개가 생성되었습니다.",
        "summary": "\n".join(summary_messages),
        "saved_count": total_saved
    }


@router.post("/generate-from-r2", summary="R2 클라우드 문서 기반 일정 생성")
async def generate_schedule_from_r2(
    prefix: str = "schedule-data-output/",
    db: Session = Depends(get_db),
    current_emp: Employee = Depends(get_current_employee)
):
    """R2에 있는 문서들을 읽어서 일정 생성 및 저장"""
    try:
        # R2에서 파일 목록 조회
        r2_files = list_r2_objects(prefix)
        
        if not r2_files:
            raise HTTPException(status_code=404, detail=f"R2 '{prefix}' 폴더에 파일이 없습니다.")
            
        tasks = db.execute(text("SELECT task_id, task_name FROM task")).mappings().all()
        task_list = "\n".join([f"- {t['task_id']}: {t['task_name']}" for t in tasks])

        orders = db.execute(text("SELECT order_id, order_num, product_name FROM orders")).mappings().all()
        order_list = "\n".join([f"- {o['order_id']}: {o['order_num']} ({o['product_name']})" for o in orders])

        employees = db.execute(text("SELECT emp_id, login_id, emp_name, emp_role FROM employees")).mappings().all()
        employee_list = "\n".join([f"- {e['emp_id']}: {e['emp_name']} ({e['emp_role']})" for e in employees])

        equipments = db.execute(text("SELECT eq_id, eq_name FROM equipments")).mappings().all()
        equipment_list = "\n".join([f"- {eq['eq_id']}: {eq['eq_name']}" for eq in equipments])

        valid_tasks = {t['task_id'] for t in tasks}

        # Map name/login_id/emp_id to emp_id
        employee_map = {}
        for e in employees:
            emp_id = e['emp_id']
            emp_name = e['emp_name']
            login_id = e['login_id']
            employee_map[emp_id.lower().strip()] = emp_id
            employee_map[emp_name.lower().strip()] = emp_id
            if login_id:
                employee_map[login_id.lower().strip()] = emp_id

        # Map name/eq_id to eq_id
        equipment_map = {}
        for eq in equipments:
            eq_id = eq['eq_id']
            eq_name = eq['eq_name']
            equipment_map[eq_id.lower().strip()] = eq_id
            equipment_map[eq_name.lower().strip()] = eq_id

        # Map order_id/order_num to order_id
        order_map = {}
        for o in orders:
            ord_id = o['order_id']
            ord_num = o['order_num']
            order_map[ord_id.lower().strip()] = ord_id
            order_map[ord_num.lower().strip()] = ord_id
        
        all_text = ""
        direct_parsed_schedules = 0
        direct_parsed_any = False
        
        # R2에서 각 파일 다운로드 및 처리
        for file_info in r2_files:
            try:
                r2_key = file_info["key"]
                filename = file_info["file_name"]
                
                # R2에서 파일 다운로드
                file_bytes = download_file_from_r2(r2_key)
                
                # 파일 파싱
                file_text = extract_text_from_bytes(file_bytes, filename)
                
                # Check if structured CSV
                if is_structured_schedule_csv(file_text):
                    print(f"ℹ️ Detected structured CSV '{filename}' in R2. Parsing directly.")
                    direct_count = parse_schedule_csv_directly(
                        file_text,
                        db=db,
                        employee_map=employee_map,
                        equipment_map=equipment_map,
                        order_map=order_map,
                        valid_tasks=valid_tasks
                    )
                    direct_parsed_schedules += direct_count
                    direct_parsed_any = True
                else:
                    print(f"ℹ️ Skipping non-schedule file: {filename}")
                    continue
                
            except Exception as e:
                print(f"⚠️ 파일 처리 오류 ({file_info['file_name']}): {str(e)}")
                continue

        llm_parsed_schedules = 0
        saved_ids = []
        combined_result = {
            "orders": [],
            "schedules": [],
            "required_equipments": [],
            "summaries": []
        }

        if all_text.strip():
            # Chunking
            all_chunks = chunk_document_text(all_text, chunk_size=100)
            print(f"ℹ️ Split document into {len(all_chunks)} chunks for LLM processing.")

            for idx, chunk in enumerate(all_chunks):
                # GPT 프롬프트
                prompt = f"""다음은 제조 공장의 생산 관련 문서들입니다. 이 문서들을 분석하여 생산 관련 데이터(주문, 일정, 작업자 배정, 필요 장비)를 JSON 형식으로 생성해주세요.

문서 내용:
{chunk}

사용 가능한 task 목록 (반드시 아래 task_id 중에서 선택):
{task_list}

사용 가능한 기존 order 목록 (문서에 있으면서 기존 목록에 있으면 해당 order_id를 사용하고, 없으면 새로운 order_id를 생성하여 orders 목록에 추가하세요):
{order_list}

사용 가능한 직원 목록 (반드시 아래 emp_id 또는 직원 이름 중에서 선택하여 일정에 배정):
{employee_list}

사용 가능한 장비 목록 (반드시 아래 eq_id 또는 장비 이름 중에서 선택하여 작업에 매핑):
{equipment_list}

다음 형식으로 응답해주세요:
{{
  "orders": [
    {{
      "order_id": "ord_001", // 기존 order_id 또는 신규 생성된 ID (예: ord_new_001)
      "order_num": "PO001",
      "product_name": "DRAM-8G",
      "order_count": 2500,
      "due_date": "2026-07-31",
      "order_status": "COMPLETED" // COMPLETED, IN_PROGRESS, PENDING 등
    }}
  ],
  "schedules": [
    {{
      "task_id": "tsk_001",
      "order_id": "ord_001",
      "start_date": "2026-07-01 09:00:00",
      "end_date": "2026-07-02 18:00:00",
      "factory": "A동", // 공장동 (예: A동, B동 등)
      "workers": ["emp001", "emp002"] // 해당 일정에 배정될 직원 ID 또는 직원 이름 목록 (최소 1명 이상 배정)
    }}
  ],
  "required_equipments": [
    {{
      "task_id": "tsk_001",
      "eq_id": "eq_001" // 해당 task에 필요한 장비 ID 또는 장비 이름
    }}
  ],
  "summary": "일정 요약 설명"
}}

주의사항:
1. schedules의 order_id는 반드시 orders 목록에 정의되어 있거나 기존 order 목록에 있는 ID/주문번호여야 합니다.
2. schedules의 task_id는 반드시 사용 가능한 task 목록에 있는 ID여야 합니다.
3. schedules의 workers는 반드시 사용 가능한 직원 목록의 emp_id 또는 직원 이름에서 선택해야 합니다. (최소 1명 이상 배정해야 일정 조회 쿼리에서 누락되지 않습니다)
4. required_equipments의 task_id 및 eq_id는 각각 제공된 task 목록과 장비 목록에 있는 ID/이름이어야 합니다.
5. JSON만 응답하고 다른 텍스트는 포함하지 마세요."""

                # GPT API 호출
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.3,
                )

                result_text = response.choices[0].message.content.strip()
                result_text = result_text.replace("```json", "").replace("```", "").strip()

                try:
                    chunk_result = json.loads(result_text)
                    combined_result["orders"].extend(chunk_result.get("orders", []))
                    combined_result["schedules"].extend(chunk_result.get("schedules", []))
                    combined_result["required_equipments"].extend(chunk_result.get("required_equipments", []))
                    if chunk_result.get("summary"):
                        combined_result["summaries"].append(chunk_result["summary"])
                except Exception as e:
                    print(f"❌ Error processing chunk {idx}: {str(e)}")
                    continue

            # Deduplicate orders and required equipments in Python
            unique_orders = []
            seen_order_ids = set()
            for o in combined_result["orders"]:
                oid = o.get("order_id")
                if oid and oid not in seen_order_ids:
                    seen_order_ids.add(oid)
                    unique_orders.append(o)

            unique_required = []
            seen_req = set()
            for r in combined_result["required_equipments"]:
                t_id = r.get("task_id")
                eq_val = r.get("eq_id")
                if t_id and eq_val:
                    key = (t_id, str(eq_val).lower().strip())
                    if key not in seen_req:
                        seen_req.add(key)
                        unique_required.append(r)

            # 1. Insert/Update Orders
            resolved_orders = {}
            for o in unique_orders:
                llm_order_id = o.get("order_id")
                order_num = o.get("order_num", "").strip()
                if not llm_order_id:
                    continue
                    
                num_key = order_num.lower().strip()
                id_key = llm_order_id.lower().strip()
                
                # Resolve to existing order_id if order_num or order_id matches
                if num_key in order_map:
                    correct_id = order_map[num_key]
                elif id_key in order_map:
                    correct_id = order_map[id_key]
                else:
                    correct_id = llm_order_id
                    order_map[id_key] = correct_id
                    if order_num:
                        order_map[num_key] = correct_id
                        
                resolved_orders[llm_order_id] = correct_id
                
                try:
                    with db.begin_nested():
                        db.execute(
                            text("""
                                INSERT INTO orders (order_id, order_num, product_name, order_count, due_date, order_status)
                                VALUES (:order_id, :order_num, :product_name, :order_count, :due_date, :order_status)
                                ON CONFLICT (order_id) DO UPDATE SET
                                    order_num = EXCLUDED.order_num,
                                    product_name = EXCLUDED.product_name,
                                    order_count = EXCLUDED.order_count,
                                    due_date = EXCLUDED.due_date,
                                    order_status = EXCLUDED.order_status
                            """),
                            {
                                "order_id": correct_id,
                                "order_num": order_num or "UNKNOWN",
                                "product_name": o.get("product_name", "UNKNOWN"),
                                "order_count": int(o.get("order_count", 0)),
                                "due_date": o.get("due_date", "2026-07-01"),
                                "order_status": o.get("order_status", "PENDING"),
                            }
                        )
                except Exception as e:
                    print(f"❌ Order insert/update error: {str(e)}")
                    continue

            # 2. Insert Required Equipments (with mathematical symbol resolution)
            for req in unique_required:
                task_id = req.get("task_id")
                eq_val = req.get("eq_id")
                if not task_id or not eq_val:
                    continue
                if task_id not in valid_tasks:
                    print(f"⚠️ Invalid task_id ({task_id}) for required_equipments, skipping.")
                    continue
                eq_key = str(eq_val).lower().strip()
                eq_id = None
                if eq_key in equipment_map:
                    eq_id = equipment_map[eq_key]
                else:
                    m = re.match(r'^장비\s*(\d+)$', str(eq_val).strip())
                    if m:
                        candidate = f"EQ{int(m.group(1)):03d}"
                        if candidate.lower() in equipment_map:
                            eq_id = equipment_map[candidate.lower()]
                    else:
                        m2 = re.match(r'^(?:[eE][qQ])?\s*(\d+)$', str(eq_val).strip())
                        if m2:
                            candidate = f"EQ{int(m2.group(1)):03d}"
                            if candidate.lower() in equipment_map:
                                eq_id = equipment_map[candidate.lower()]
                if not eq_id:
                    print(f"⚠️ Equipment '{eq_val}' not resolved, skipping required_equipment mapping.")
                    continue
                try:
                    with db.begin_nested():
                        db.execute(
                            text("""
                                INSERT INTO required_equipments (task_id, eq_id)
                                VALUES (:task_id, :eq_id)
                                ON CONFLICT (task_id, eq_id) DO NOTHING
                            """),
                            {
                                "task_id": task_id,
                                "eq_id": eq_id
                            }
                        )
                except Exception as e:
                    print(f"❌ Required equipment insert error: {str(e)}")
                    continue

            # 3. Insert/Update Schedules
            saved = []
            for s in combined_result["schedules"]:
                task_id = s.get("task_id")
                order_val = s.get("order_id")
                if not task_id or not order_val:
                    continue
                if task_id not in valid_tasks:
                    print(f"⚠️ Invalid task_id ({task_id}) for schedule, skipping.")
                    continue
                
                # Resolve order_id
                if order_val in resolved_orders:
                    order_id = resolved_orders[order_val]
                else:
                    order_key = str(order_val).lower().strip()
                    if order_key in order_map:
                        order_id = order_map[order_key]
                    else:
                        print(f"⚠️ Order_id ({order_val}) not found in orders database or generated list, skipping schedule.")
                        continue

                # Check if schedule already exists for this task and order to prevent duplicates
                existing = db.execute(
                    text("SELECT id FROM schedules WHERE task_id = :task_id AND order_id = :order_id"),
                    {"task_id": task_id, "order_id": order_id}
                ).mappings().first()

                if existing:
                    schedule_id = existing["id"]
                    try:
                        with db.begin_nested():
                            db.execute(
                                text("""
                                    UPDATE schedules 
                                    SET start_date = :start_date, end_date = :end_date, factory = :factory
                                    WHERE id = :id AND task_id = :task_id AND order_id = :order_id
                                """),
                                {
                                    "id": schedule_id,
                                    "task_id": task_id,
                                    "order_id": order_id,
                                    "start_date": s["start_date"],
                                    "end_date": s["end_date"],
                                    "factory": s.get("factory", "A동"),
                                }
                            )
                            # Clean up old assignments for this schedule to prevent duplicates
                            db.execute(
                                text("""
                                    DELETE FROM schedule_assignments 
                                    WHERE id = :id AND task_id = :task_id AND order_id = :order_id
                                """),
                                {
                                    "id": schedule_id,
                                    "task_id": task_id,
                                    "order_id": order_id
                                }
                            )
                    except Exception as e:
                        print(f"❌ Schedule update error: {str(e)}")
                        continue
                else:
                    schedule_id = f"sch_{uuid.uuid4().hex[:8]}"
                    try:
                        with db.begin_nested():
                            db.execute(
                                text("""
                                    INSERT INTO schedules (id, task_id, order_id, start_date, end_date, factory)
                                    VALUES (:id, :task_id, :order_id, :start_date, :end_date, :factory)
                                """),
                                {
                                    "id": schedule_id,
                                    "task_id": task_id,
                                    "order_id": order_id,
                                    "start_date": s["start_date"],
                                    "end_date": s["end_date"],
                                    "factory": s.get("factory", "A동"),
                                }
                            )
                    except Exception as e:
                        print(f"❌ DB 저장 오류: {str(e)}")
                        continue

                # Map workers to valid employee IDs
                workers = []
                for w in s.get("workers", []):
                    w_key = str(w).lower().strip()
                    if w_key in employee_map:
                        workers.append(employee_map[w_key])
                    else:
                        print(f"⚠️ Worker value '{w}' not resolved to an employee.")
                
                # De-duplicate workers in Python
                workers = list(set(workers))
                saved.append((schedule_id, task_id, order_id, workers))

            # 4. Insert Schedule Assignments
            for schedule_id, task_id, order_id, workers in saved:
                has_at_least_one_worker = False
                for worker_id in workers:
                    try:
                        with db.begin_nested():
                            db.execute(
                                text("""
                                    INSERT INTO schedule_assignments (id, user_id, task_id, order_id)
                                    VALUES (:id, :user_id, :task_id, :order_id)
                                    ON CONFLICT (id, user_id, task_id, order_id) DO NOTHING
                                """),
                                {
                                    "id": schedule_id,
                                    "user_id": worker_id,
                                    "task_id": task_id,
                                    "order_id": order_id
                                }
                            )
                        has_at_least_one_worker = True
                    except Exception as e:
                        print(f"❌ Assignment insert error: {str(e)}")
                        continue
                if has_at_least_one_worker:
                    llm_parsed_schedules += 1
                    saved_ids.append(schedule_id)

            db.commit()

        total_saved = direct_parsed_schedules + llm_parsed_schedules
        summary_messages = []
        if direct_parsed_any:
            summary_messages.append(f"구조화된 CSV 파일에서 {direct_parsed_schedules}개 일정을 직접 파싱하여 저장했습니다.")
        if all_text.strip():
            summary_messages.append("\n".join(combined_result["summaries"]))

        return {
            "message": f"✅ R2에서 {len(r2_files)}개 파일을 읽어 일정 {total_saved}개가 생성되었습니다.",
            "summary": "\n".join(summary_messages),
            "file_count": len(r2_files),
            "saved_count": total_saved,
            "saved_ids": saved_ids
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"일정 생성 실패: {str(e)}")